"""
大厨技能 - Excel 饮食总结报告生成器
读取饮食数据，生成格式化的多Sheet Excel报告。
"""
import json
import os
from datetime import datetime, timedelta
from collections import Counter, defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    exit(1)

# 数据文件路径
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

# 样式定义
HEADER_FONT = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="微软雅黑", bold=True, size=11, color="2F5496")
NORMAL_FONT = Font(name="微软雅黑", size=10)
WARN_FONT = Font(name="微软雅黑", size=10, color="FF0000")
GOOD_FONT = Font(name="微软雅黑", size=10, color="008000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 分类颜色
CATEGORY_FILLS = {
    "蔬菜": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "肉类": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "海鲜": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "豆制品": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "主食": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    "蛋类": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    "乳制品": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "水果": PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
}


def load_json(filename):
    """加载JSON数据文件"""
    filepath = os.path.join(DATA_DIR, filename)
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def apply_header_style(ws, row, max_col):
    """给表头行应用样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def apply_cell_style(cell, font=None, align=None, fill=None):
    """应用单元格样式"""
    cell.font = font or NORMAL_FONT
    cell.alignment = align or CENTER_ALIGN
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill


def auto_width(ws, max_col, min_width=10, max_width=35):
    """自动调整列宽"""
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    # 中文字符宽度约等于2个英文字符
                    val = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_len = max(max_len, char_len + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len, max_width)


def generate_report(start_date, end_date, output_path):
    """
    生成饮食总结报告
    start_date: 开始日期 "YYYY-MM-DD"
    end_date: 结束日期 "YYYY-MM-DD"
    output_path: 输出Excel文件路径
    """
    meal_log = load_json("meal_log.json")
    recipes = load_json("recipes.json")
    restrictions = load_json("restrictions.json")

    records = meal_log.get("records", [])
    recipe_dict = {r["id"]: r for r in recipes.get("recipes", [])}

    # 筛选日期范围内的记录
    filtered = [r for r in records if start_date <= r["date"] <= end_date]
    if not filtered:
        print(f"日期范围 {start_date} ~ {end_date} 内没有饮食记录")
        return None

    # 按日期和餐次分组
    daily_summary = defaultdict(lambda: {"早餐": [], "午餐": [], "晚餐": [], "加餐": [], "total_cal": 0,
                                          "total_protein": 0, "total_fat": 0, "total_carbs": 0})
    for r in filtered:
        d = daily_summary[r["date"]]
        d[r["meal_type"]] = r.get("dishes", [])
        d["total_cal"] += r.get("total_calories", 0)

    sorted_dates = sorted(daily_summary.keys())

    # 获取营养目标
    goals = restrictions.get("health_goals", {})
    target_cal = goals.get("target_calories", 2000)
    target_protein = goals.get("target_protein", 60)
    target_fat = goals.get("target_fat", 55)
    target_carbs = goals.get("target_carbs", 300)

    # 统计菜品频次
    dish_counter = Counter()
    all_ingredients = Counter()
    for r in filtered:
        for dish in r.get("dishes", []):
            dish_counter[dish["name"]] += dish.get("portions", 1)
            # 查菜谱获取食材
            rid = dish.get("recipe_id")
            if rid and rid in recipe_dict:
                for ing in recipe_dict[rid].get("ingredients", []):
                    all_ingredients[ing["name"]] += ing.get("amount", 0)

    wb = Workbook()

    # ===== Sheet 1: 饮食概览 =====
    ws1 = wb.active
    ws1.title = "饮食概览"
    ws1.merge_cells("A1:H1")
    ws1.cell(row=1, column=1, value=f"饮食概览（{start_date} ~ {end_date}）").font = TITLE_FONT

    # 每日汇总表
    headers1 = ["日期", "早餐", "午餐", "晚餐", "加餐", "热量(kcal)", "达标", "评价"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=3, column=col, value=h)
    apply_header_style(ws1, 3, len(headers1))

    total_days = len(sorted_dates)
    total_cal_all = 0
    for i, date in enumerate(sorted_dates):
        row = 4 + i
        d = daily_summary[date]
        ws1.cell(row=row, column=1, value=date)
        ws1.cell(row=row, column=2, value="; ".join(f"{x['name']}" for x in d["早餐"]))
        ws1.cell(row=row, column=3, value="; ".join(f"{x['name']}" for x in d["午餐"]))
        ws1.cell(row=row, column=4, value="; ".join(f"{x['name']}" for x in d["晚餐"]))
        ws1.cell(row=row, column=5, value="; ".join(f"{x['name']}" for x in d["加餐"]))
        ws1.cell(row=row, column=6, value=d["total_cal"])

        ratio = d["total_cal"] / target_cal if target_cal else 0
        if 0.9 <= ratio <= 1.1:
            ws1.cell(row=row, column=7, value="✓ 达标")
            ws1.cell(row=row, column=8, value="不错！热量控制得当")
        elif ratio < 0.9:
            ws1.cell(row=row, column=7, value="↓ 偏低")
            ws1.cell(row=row, column=8, value="摄入偏低，注意营养")
        else:
            ws1.cell(row=row, column=7, value="↑ 偏高")
            ws1.cell(row=row, column=8, value="摄入偏高，适当控制")

        for col in range(1, len(headers1) + 1):
            apply_cell_style(ws1.cell(row=row, column=col), align=LEFT_ALIGN if col in [2,3,4,5,8] else CENTER_ALIGN)
        total_cal_all += d["total_cal"]

    # 汇总行
    summary_row = 4 + total_days
    ws1.cell(row=summary_row, column=1, value="汇总").font = SUBTITLE_FONT
    avg_cal = total_cal_all / total_days if total_days > 0 else 0
    ws1.cell(row=summary_row, column=6, value=f"日均: {avg_cal:.0f}").font = SUBTITLE_FONT
    ws1.cell(row=summary_row, column=8, value=f"目标: {target_cal} kcal/天").font = SUBTITLE_FONT
    for col in range(1, len(headers1) + 1):
        apply_cell_style(ws1.cell(row=summary_row, column=col),
                        font=SUBTITLE_FONT, fill=PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"))

    auto_width(ws1, len(headers1))

    # 添加图表
    chart = BarChart()
    chart.type = "col"
    chart.title = "每日热量摄入趋势"
    chart.y_axis.title = "热量 (kcal)"
    chart.x_axis.title = "日期"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    data_ref = Reference(ws1, min_col=6, min_row=3, max_row=3 + total_days)
    cats_ref = Reference(ws1, min_col=1, min_row=4, max_row=3 + total_days)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "4472C4"
    ws1.add_chart(chart, "J3")

    # ===== Sheet 2: 营养分析 =====
    ws2 = wb.create_sheet("营养分析")
    ws2.merge_cells("A1:G1")
    ws2.cell(row=1, column=1, value=f"营养分析（{start_date} ~ {end_date}）").font = TITLE_FONT
    ws2.cell(row=2, column=1, value=f"营养目标: 热量{target_cal}kcal | 蛋白质{target_protein}g | 脂肪{target_fat}g | 碳水{target_carbs}g").font = NORMAL_FONT

    headers2 = ["日期", "热量(kcal)", "偏离目标", "蛋白质(g)", "脂肪(g)", "碳水(g)", "评价"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=col, value=h)
    apply_header_style(ws2, 4, len(headers2))

    for i, date in enumerate(sorted_dates):
        row = 5 + i
        d = daily_summary[date]
        cal = d["total_cal"]
        diff = cal - target_cal
        diff_pct = (diff / target_cal * 100) if target_cal else 0
        ws2.cell(row=row, column=1, value=date)
        ws2.cell(row=row, column=2, value=cal)
        ws2.cell(row=row, column=3, value=f"{diff:+.0f} ({diff_pct:+.1f}%)")
        ws2.cell(row=row, column=4, value=round(cal * 0.15 / 4, 0))
        ws2.cell(row=row, column=5, value=round(cal * 0.25 / 9, 0))
        ws2.cell(row=row, column=6, value=round(cal * 0.60 / 4, 0))

        if abs(diff_pct) <= 10:
            ws2.cell(row=row, column=7, value="正常 ✓")
            ws2.cell(row=row, column=7).font = GOOD_FONT
        elif diff_pct > 10:
            ws2.cell(row=row, column=7, value="偏高 ↑")
            ws2.cell(row=row, column=7).font = WARN_FONT
        else:
            ws2.cell(row=row, column=7, value="偏低 ↓")
            ws2.cell(row=row, column=7).font = WARN_FONT

        for col in range(1, len(headers2) + 1):
            apply_cell_style(ws2.cell(row=row, column=col))

    auto_width(ws2, len(headers2))

    # 饼图：三大营养素占比
    pie = PieChart()
    pie.title = "三大营养素平均占比"
    pie.style = 10
    pie.width = 16
    pie.height = 12
    avg_protein = sum(ws2.cell(row=5 + i, column=4).value or 0 for i in range(total_days)) / total_days * 4
    avg_fat = sum(ws2.cell(row=5 + i, column=5).value or 0 for i in range(total_days)) / total_days * 9
    avg_carbs = sum(ws2.cell(row=5 + i, column=6).value or 0 for i in range(total_days)) / total_days * 4
    # 添加饼图数据
    pie_data = [["营养素", "热量占比"],
                ["蛋白质", round(avg_protein)],
                ["脂肪", round(avg_fat)],
                ["碳水", round(avg_carbs)]]
    for i, item in enumerate(pie_data):
        if i == 0:
            continue
        ws2.cell(row=5 + total_days + i, column=1, value=item[0])
        ws2.cell(row=5 + total_days + i, column=2, value=item[1])
    data_ref = Reference(ws2, min_col=2, min_row=5 + total_days + 1, max_row=5 + total_days + 3)
    cats_ref = Reference(ws2, min_col=1, min_row=5 + total_days + 1, max_row=5 + total_days + 3)
    pie.add_data(data_ref, titles_from_data=False)
    pie.set_categories(cats_ref)
    colors = ["FF6B6B", "FFD93D", "6BCB77"]
    for idx, color in enumerate(colors):
        pie.series[0].data_points[idx].graphicalProperties.solidFill = color
    ws2.add_chart(pie, "J3")

    # ===== Sheet 3: 菜品排行 =====
    ws3 = wb.create_sheet("菜品排行")
    ws3.merge_cells("A1:E1")
    ws3.cell(row=1, column=1, value="菜品食用频率排行").font = TITLE_FONT

    headers3 = ["排名", "菜品名称", "食用次数", "场景", "备注"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=col, value=h)
    apply_header_style(ws3, 3, len(headers3))

    top_dishes = dish_counter.most_common(20)
    for i, (name, count) in enumerate(top_dishes):
        row = 4 + i
        ws3.cell(row=row, column=1, value=f"#{i + 1}")
        ws3.cell(row=row, column=2, value=name)
        ws3.cell(row=row, column=3, value=count)
        # 查找菜谱信息
        scene = ""
        notes = ""
        for rid, rcp in recipe_dict.items():
            if rcp["name"] == name:
                scene = ", ".join(rcp.get("scene", []))
                notes = rcp.get("notes", "")
                break
        ws3.cell(row=row, column=4, value=scene)
        ws3.cell(row=row, column=5, value=notes)
        for col in range(1, len(headers3) + 1):
            apply_cell_style(ws3.cell(row=row, column=col))
        if i < 3:
            ws3.cell(row=row, column=1).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    auto_width(ws3, len(headers3))

    # 柱状图：菜品频次
    chart3 = BarChart()
    chart3.type = "bar"
    chart3.title = "菜品食用频次 TOP 10"
    chart3.style = 10
    chart3.width = 22
    chart3.height = 14
    data_ref3 = Reference(ws3, min_col=3, min_row=3, max_row=3 + min(10, len(top_dishes)))
    cats_ref3 = Reference(ws3, min_col=2, min_row=4, max_row=3 + min(10, len(top_dishes)))
    chart3.add_data(data_ref3, titles_from_data=True)
    chart3.set_categories(cats_ref3)
    chart3.series[0].graphicalProperties.solidFill = "70AD47"
    ws3.add_chart(chart3, "G3")

    # ===== Sheet 4: 食材消耗 =====
    ws4 = wb.create_sheet("食材消耗")
    ws4.merge_cells("A1:E1")
    ws4.cell(row=1, column=1, value="食材消耗统计").font = TITLE_FONT

    headers4 = ["排名", "食材名称", "消耗量", "分类", "评价"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=col, value=h)
    apply_header_style(ws4, 3, len(headers4))

    # 简单的食材分类映射
    ingredient_cats = {
        "鸡胸肉": "肉类", "猪肉": "肉类", "牛肉": "肉类", "羊肉": "肉类",
        "鸡蛋": "蛋类", "鸭蛋": "蛋类",
        "豆腐": "豆制品", "豆干": "豆制品",
        "青菜": "蔬菜", "白菜": "蔬菜", "西兰花": "蔬菜", "胡萝卜": "蔬菜",
        "青椒": "蔬菜", "番茄": "蔬菜", "黄瓜": "蔬菜", "茄子": "蔬菜",
        "土豆": "蔬菜", "洋葱": "蔬菜", "生菜": "蔬菜", "菠菜": "蔬菜",
        "虾": "海鲜", "鱼": "海鲜", "蟹": "海鲜",
        "大米": "主食", "面条": "主食", "馒头": "主食", "面包": "主食",
        "牛奶": "乳制品", "酸奶": "乳制品",
        "苹果": "水果", "香蕉": "水果", "橙子": "水果",
    }

    top_ingredients = all_ingredients.most_common(20)
    for i, (name, amount) in enumerate(top_ingredients):
        row = 4 + i
        cat = ingredient_cats.get(name, "其他")
        ws4.cell(row=row, column=1, value=f"#{i + 1}")
        ws4.cell(row=row, column=2, value=name)
        ws4.cell(row=row, column=3, value=f"{amount:.0f}")
        ws4.cell(row=row, column=4, value=cat)

        # 评价食材多样性
        if amount >= 5:
            ws4.cell(row=row, column=5, value="高频使用 ⭐")
        elif amount >= 3:
            ws4.cell(row=row, column=5, value="常用 ✅")
        else:
            ws4.cell(row=row, column=5, value="偶尔使用")

        for col in range(1, len(headers4) + 1):
            apply_cell_style(ws4.cell(row=row, column=col),
                           fill=CATEGORY_FILLS.get(cat))

    auto_width(ws4, len(headers4))

    # ===== Sheet 5: 改善建议 =====
    ws5 = wb.create_sheet("改善建议")
    ws5.merge_cells("A1:C1")
    ws5.cell(row=1, column=1, value="饮食改善建议").font = TITLE_FONT

    tips = []
    tips.append(["问题/发现", "建议", "优先级"])
    tips.append(["", "", ""])

    # 分析建议
    over_days = sum(1 for d in daily_summary.values() if d["total_cal"] > target_cal * 1.1)
    under_days = sum(1 for d in daily_summary.values() if d["total_cal"] < target_cal * 0.9)

    if over_days > total_days * 0.3:
        tips.append(["热量超标天数较多", f"{over_days}/{total_days}天超出目标热量，建议减少高油高糖食物，增加蔬菜比例", "高"])
    if under_days > total_days * 0.3:
        tips.append(["热量偏低天数较多", f"{under_days}/{total_days}天低于目标热量，建议适当增加主食或蛋白质摄入", "高"])

    # 食材多样性分析
    unique_ingredients = len(all_ingredients)
    if unique_ingredients < 15:
        tips.append(["食材多样性不足", f"本周期仅摄入{unique_ingredients}种不同食材，建议每周摄入25种以上不同食材，多尝试新菜品", "中"])
    elif unique_ingredients < 25:
        tips.append(["食材多样性可提升", f"本周期摄入{unique_ingredients}种食材，已接近推荐标准，可再增加蔬菜水果种类", "低"])

    # 菜品重复度
    if dish_counter and dish_counter.most_common(1)[0][1] > total_days:
        top_dish = dish_counter.most_common(1)[0]
        tips.append(["菜品重复度较高", f"「{top_dish[0]}」出现{top_dish[1]}次，建议轮换菜品以保持营养均衡和新鲜感", "中"])

    # 早餐频率
    breakfast_count = sum(1 for d in daily_summary.values() if d["早餐"])
    if breakfast_count < total_days:
        tips.append([f"有{total_days - breakfast_count}天未记录早餐", "规律早餐有助于维持代谢水平和新的一天精力", "高"])

    # 如果没有具体建议
    if len(tips) <= 2:
        tips.append(["整体表现良好", "饮食结构比较均衡，继续保持！", "低"])

    for col, h in enumerate(tips[0], 1):
        ws5.cell(row=3, column=col, value=h)
    apply_header_style(ws5, 3, 3)

    for i, tip in enumerate(tips[1:], 4):
        for col, val in enumerate(tip, 1):
            cell = ws5.cell(row=i, column=col, value=val)
            apply_cell_style(cell, align=LEFT_ALIGN)
            if col == 3 and val in ["高", "中"]:
                cell.font = WARN_FONT if val == "高" else Font(name="微软雅黑", size=10, color="FF8C00")

    auto_width(ws5, 3)
    ws5.column_dimensions["B"].width = 60

    # 冻结首行
    for ws in [ws1, ws2, ws3, ws4]:
        ws.freeze_panes = "A4"
    ws5.freeze_panes = "A4"

    # 保存
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # 独立运行时：生成最近7天的报告
    end = datetime.now()
    start = end - timedelta(days=6)
    output = os.path.join(os.path.dirname(os.path.dirname(DATA_DIR)),
                          f"饮食报告_{start.strftime('%m%d')}_{end.strftime('%m%d')}.xlsx")
    result = generate_report(start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"), output)
    if result:
        print(f"报告已生成: {result}")
    else:
        print("未能生成报告，请先使用大厨技能记录饮食数据。")
